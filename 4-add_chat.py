from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Lê a pasta de trabalho e planilha
wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatório"]

# Referências das linhas e colunas
min_column = wb.active.min_column
min_row = wb.active.min_row
max_column = wb.active.max_column
max_row = wb.active.max_row
# print(min_column, min_row)
# print(max_column, max_row)

# Adicionando dados e categorias no gráfico
barChart = BarChart()

data = Reference(
    sheet,
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row,
    max_row=max_row
)

categories = Reference(
    sheet,
    min_col=min_column,
    max_col=min_column,
    min_row=min_row + 1,
    max_row=max_row
)

barChart.add_data(data, titles_from_data=True)
barChart.set_categories(categories)

# Criando o gráfico
sheet.add_chart(barChart, "B10")
barChart.title = "Vendas por Fabricante"
barChart.style = 2

# Salvando o Workbook
wb.save("data/barchart.xlsx")